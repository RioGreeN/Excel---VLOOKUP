import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from tkinter import Tk
from tkinter.filedialog import askdirectory

def select_folder():
    """
    弹出文件夹选择对话框，让用户选择包含 Excel 文件的文件夹。
    :return: 用户选择的文件夹路径
    """
    Tk().withdraw()  # 隐藏 Tkinter 根窗口
    folder_path = askdirectory(title="选择包含 Excel 文件的文件夹")
    if not folder_path:
        raise ValueError("未选择任何文件夹，程序退出。")
    return folder_path

def merge_excel_files_with_images(folder_path, key_column, output_file):
    """
    合并文件夹内多个 Excel 文件，按照指定列进行合并，并保留 Excel 文件中的图片。

    :param folder_path: 文件夹路径，包含所有待合并的 Excel 文件
    :param key_column: 用于合并的列名
    :param output_file: 输出合并结果的 Excel 文件路径
    """
    # 用于存储所有 Excel 数据的列表
    dataframes = []
    images_dict = {}  # 用于存储每个文件的图片信息

    # 遍历文件夹中的 Excel 文件
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):  # 只处理 .xlsx 文件
            file_path = os.path.join(folder_path, file_name)
            print(f"正在处理文件: {file_path}")

            # 使用 pandas 读取表格数据
            df = pd.read_excel(file_path)

            # 确保关键列存在
            if key_column not in df.columns:
                raise ValueError(f"文件 {file_name} 中缺少关键列 '{key_column}'")

            dataframes.append(df)  # 添加数据到列表中

            # 使用 openpyxl 读取图片及其位置
            wb = load_workbook(file_path)
            ws = wb.active
            images = []
            for image in ws._images:
                images.append((image, image.anchor))  # 存储图片和锚点位置
            images_dict[file_name] = images

    # 使用 pandas 按关键列合并所有数据
    print("正在合并数据...")
    merged_data = dataframes[0]
    for df in dataframes[1:]:
        merged_data = pd.merge(merged_data, df, on=key_column, how='outer')  # 使用 outer 保留所有数据

    # 创建一个新的工作簿保存合并结果
    print(f"正在将合并结果保存到: {output_file}")
    wb = Workbook()
    ws = wb.active

    # 写入合并后的数据到新工作簿
    ws.append(list(merged_data.columns))  # 写入标题
    for _, row in merged_data.iterrows():
        ws.append(list(row))

    # 插入图片到新工作簿
    print("正在保留图片...")
    for file_name, images in images_dict.items():
        for img, anchor in images:
            # 将图片插入到对应的单元格
            ws.add_image(img, anchor)

    # 保存结果
    wb.save(output_file)
    print("合并完成！")

if __name__ == "__main__":
    # 手动选择文件夹
    try:
        folder_path = select_folder()
        key_column = input("请输入关键字：（默认为SN）")
        if key_column == '':
            key_column = "SN"  # 替换为用于合并的列名
        output_file = "merged_data_with_images.xlsx"  # 合并结果保存的文件名

        merge_excel_files_with_images(folder_path, key_column, output_file)
    except ValueError as e:
        print(e)
