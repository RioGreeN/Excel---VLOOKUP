import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl import Workbook

def merge_excel_files_with_images(folder_path, key_column, output_file):
    """
    合并文件夹内多个 Excel 文件，按照指定的列进行合并，并保留 Excel 文件中的图片。

    :param folder_path: 文件夹路径，包含所有待合并的 Excel 文件
    :param key_column: 用于合并的列名
    :param output_file: 输出合并结果的 Excel 文件路径
    """
    # 用于存储所有 Excel 数据的列表
    dataframes = []
    images_dict = {}  # 用于存储每个文件的图片信息
    
    # 遍历文件夹内的 Excel 文件
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):  # 只处理 .xlsx 文件
            file_path = os.path.join(folder_path, file_name)
            print(f"正在处理文件: {file_path}")
            
            # 使用 pandas 读取表格数据
            df = pd.read_excel(file_path)
            
            # 确保关键列存在
            if key_column not in df.columns:
                raise ValueError(f"文件 {file_name} 中缺少关键列 '{key_column}'")
            
            dataframes.append(df)  # 添加数据到列表中

            # 使用 openpyxl 读取图片
            wb = load_workbook(file_path)
            ws = wb.active
            images = []
            for drawing in ws._images:  # 获取所有嵌入图片
                images.append(drawing)
            images_dict[file_name] = images  # 存储图片信息
    
    # 使用 pandas 按关键列合并所有数据
    print("正在合并数据...")
    merged_data = dataframes[0]
    for df in dataframes[1:]:
        merged_data = pd.merge(merged_data, df, on=key_column, how='outer')  # 使用 outer 保留所有数据

    # 创建一个新的工作簿保存合并结果
    print(f"正在将合并结果保存到: {output_file}")
    wb = Workbook()
    ws = wb.active

    # 写入合并后的数据
    for r_idx, row in merged_data.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    
    # 将图片插入到新工作簿中
    print("正在保留图片...")
    current_row = 2  # 数据从第 2 行开始
    for file_name, images in images_dict.items():
        for img in images:
            img.anchor = f"A{current_row}"  # 假设图片插入到第 A 列
            ws.add_image(img)
            current_row += 5  # 假设每张图片占用 5 行

    # 保存结果
    wb.save(output_file)
    print("合并完成！")



folder_path = 'C:\\Users\\MrGreed\\Desktop\\新建文件夹'  # 替换为包含 Excel 文件的文件夹路径
key_column = 'SN'  # 替换为合并所依据的列名
output_file = folder_path + "\\" + "合并表格.xlsx"  # 合并结果保存的文件名



merge_excel_files_with_images(folder_path, key_column, output_file)